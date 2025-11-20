# ref. v136
# pipenv install openpyxl

import os
import openpyxl


os.system('cls' if os.name == 'nt' else 'clear')

wb = openpyxl.load_workbook("planilla.xlsx")
print("hojas de calculo del archivo", wb.sheetnames)

hoja = wb.active
print(hoja)

wb.create_sheet("Hoja 3")
print("hojas de calculo del archivo", wb.sheetnames)

hoja3 = wb["Hoja 3"]
hoja3.title = "Nuevo titulo"
print("hojas de calculo del archivo", wb.sheetnames)


print("filas = ", hoja.max_row) # imprime 34 pero solo tengo 4 5 lineas con datos
print("columnas = ", hoja.max_column)

celda = hoja["A3"]
celda.value = "rolUsuario"
print(celda.value)


celda2 = hoja.cell(row=3, column=2)
print(celda2.value)
print(celda2.row)
print(celda2.column)
print(celda2.coordinate)

for fila in range(1, hoja.max_row + 1):
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila, column=columna)
        # print(celda)

columna = hoja["B"]
print(columna)

fila = hoja["3"]
print(fila)

hoja.append([1050, "mnunez@cecaitra.org.ar", 20])
print(hoja.rows)

hoja.delete_rows(2, 1)
hoja.delete_rows(4, 1)
# hoja.delete_cols()

wb.save("nuevo_excel.xlsx")